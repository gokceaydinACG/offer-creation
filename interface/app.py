# interface/app.py
"""
Offer Creation Tool - Main Application

Clean, modular Streamlit interface for converting supplier offers.
"""

import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st

from components import (
    render_department_selector,
    render_download_buttons,
    render_file_uploader,
    render_header,
    render_process_button,
    render_product_image_uploader,
    render_reset_button,
    render_selectable_table,
    render_success_message,
)
from processor import process_uploaded_file
from styles import get_custom_css

# Import configuration
import sys
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent))

from config import (
    MAX_FILE_SIZE_MB,
    MAX_SHEET_ROWS,
    MAX_SHEET_COLS,
    MAX_SHEETS,
    EXTREME_COLS_LIMIT,
)


def _force_availability_ints(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure availability columns are displayed as integers in the UI (no .0),
    using CEIL for any non-integer numeric values.
    """
    if df is None or df.empty:
        return df

    cols = ["Availability/Cartons", "Availability/Pieces", "Availability/Pallets"]
    for c in cols:
        if c in df.columns:
            s = pd.to_numeric(df[c], errors="coerce")
            s = np.ceil(s)
            df[c] = s.astype("Int64")  # nullable integer
    return df


def _validate_excel_file(uploaded_file) -> tuple[bool, str, dict]:
    """
    Validate Excel file size and structure.
    
    Returns:
        (is_valid, error_message, sheet_info_dict)
        sheet_info_dict = {sheet_name: {"rows": int, "cols": int}}
    """
    import openpyxl
    
    # 1. Check file size
    file_size_mb = uploaded_file.size / (1024 * 1024)
    if file_size_mb > MAX_FILE_SIZE_MB:
        return False, f"❌ File size ({file_size_mb:.1f} MB) exceeds limit ({MAX_FILE_SIZE_MB} MB). Please reduce file size.", {}
    
    # 2. Save temp file and inspect with openpyxl
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getbuffer())
            tmp_path = tmp.name
        
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        
        sheet_info = {}
        total_sheets = len(wb.sheetnames)
        
        if total_sheets > MAX_SHEETS:
            wb.close()
            Path(tmp_path).unlink(missing_ok=True)
            return False, f"❌ File has {total_sheets} sheets. Maximum {MAX_SHEETS} sheets allowed.", {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = ws.max_row
            cols = ws.max_column
            
            sheet_info[sheet_name] = {"rows": rows, "cols": cols}
        
        wb.close()
        Path(tmp_path).unlink(missing_ok=True)
        
        return True, "", sheet_info
        
    except Exception as e:
        return False, f"❌ Error reading Excel file: {str(e)}", {}


def _check_sheet_limits(sheet_info: dict, selected_sheet: str) -> tuple[bool, str]:
    """
    Check if selected sheet exceeds limits.
    
    Returns:
        (is_valid, error_message)
    """
    if selected_sheet not in sheet_info:
        return False, "❌ Selected sheet not found in file."
    
    info = sheet_info[selected_sheet]
    rows = info["rows"]
    cols = info["cols"]
    
    # Hard block extremely wide sheets
    if cols > EXTREME_COLS_LIMIT:
        return False, f"❌ Sheet has {cols:,} columns (limit: {EXTREME_COLS_LIMIT:,}). This is too wide to process. Please split the data."
    
    # Check row limit
    if rows > MAX_SHEET_ROWS:
        return False, f"❌ Sheet has {rows:,} rows (limit: {MAX_SHEET_ROWS:,}). Please filter or split the data."
    
    # Check column limit
    if cols > MAX_SHEET_COLS:
        return False, f"❌ Sheet has {cols:,} columns (limit: {MAX_SHEET_COLS:,}). Please reduce columns."
    
    return True, ""


# ============================================================================
# PAGE CONFIG
# ============================================================================
st.set_page_config(
    page_title="Offer Creator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ============================================================================
# APPLY STYLES
# ============================================================================
st.markdown(get_custom_css(), unsafe_allow_html=True)

# ============================================================================
# SESSION STATE INITIALIZATION
# ============================================================================
if "processed" not in st.session_state:
    st.session_state.processed = False
if "output_path" not in st.session_state:
    st.session_state.output_path = None
if "df" not in st.session_state:
    st.session_state.df = None
if "selected_df" not in st.session_state:
    st.session_state.selected_df = None
if "row_selected" not in st.session_state:
    st.session_state.row_selected = None
if "double_stackable" not in st.session_state:
    st.session_state.double_stackable = False
if "extract_price" not in st.session_state:
    st.session_state.extract_price = False
if "product_images" not in st.session_state:
    st.session_state.product_images = {}
if "uploaded_file_data" not in st.session_state:
    st.session_state.uploaded_file_data = None
if "dept_type" not in st.session_state:
    st.session_state.dept_type = None
if "sheet_info" not in st.session_state:
    st.session_state.sheet_info = {}
if "selected_sheet" not in st.session_state:
    st.session_state.selected_sheet = None
if "file_validated" not in st.session_state:
    st.session_state.file_validated = False

# ============================================================================
# MAIN APP FLOW
# ============================================================================
render_header()

dept_type, double_stackable, extract_price = render_department_selector()
st.session_state.double_stackable = double_stackable
st.session_state.extract_price = extract_price
st.session_state.dept_type = dept_type

if dept_type:
    uploaded_file = render_file_uploader()

    if uploaded_file:
        st.session_state.uploaded_file_data = uploaded_file
        
        # ====================================================================
        # NEW: VALIDATE FILE AND SHOW SHEET INFO
        # ====================================================================
        
        # Only validate once per file
        if not st.session_state.file_validated or st.session_state.get("last_file_name") != uploaded_file.name:
            with st.spinner("🔍 Validating file..."):
                is_valid, error_msg, sheet_info = _validate_excel_file(uploaded_file)
                
                if not is_valid:
                    st.error(error_msg)
                    st.stop()
                
                st.session_state.sheet_info = sheet_info
                st.session_state.file_validated = True
                st.session_state.last_file_name = uploaded_file.name
                st.session_state.selected_sheet = None  # Reset sheet selection
        
        # ====================================================================
        # NEW: SHEET SELECTION UI
        # ====================================================================
        
        if st.session_state.sheet_info:
            st.markdown("---")
            st.subheader("📋 Select Sheet to Process")
            
            # Display sheet info in a nice format
            sheet_options = []
            for sheet_name, info in st.session_state.sheet_info.items():
                rows = info["rows"]
                cols = info["cols"]
                
                # Check if sheet exceeds limits
                exceeds_rows = rows > MAX_SHEET_ROWS
                exceeds_cols = cols > MAX_SHEET_COLS
                exceeds_extreme = cols > EXTREME_COLS_LIMIT
                
                status = "✅"
                if exceeds_extreme:
                    status = "🚫 TOO WIDE"
                elif exceeds_rows or exceeds_cols:
                    status = "⚠️ EXCEEDS LIMIT"
                
                label = f"{status} {sheet_name} ({rows:,} rows × {cols:,} cols)"
                sheet_options.append((sheet_name, label, exceeds_rows or exceeds_cols or exceeds_extreme))
            
            # Create selectbox with formatted options
            selected_label = st.selectbox(
                "Choose a sheet:",
                options=[label for _, label, _ in sheet_options],
                index=None,
                help="Select the sheet containing your supplier offer data"
            )
            
            # Extract actual sheet name from label
            if selected_label:
                selected_sheet_name = None
                for sheet_name, label, exceeds in sheet_options:
                    if label == selected_label:
                        selected_sheet_name = sheet_name
                        break
                
                st.session_state.selected_sheet = selected_sheet_name
                
                # Show warning if limits exceeded
                is_valid, limit_error = _check_sheet_limits(st.session_state.sheet_info, selected_sheet_name)
                if not is_valid:
                    st.error(limit_error)
                    st.info("💡 **Tips to reduce file size:**\n"
                           "- Filter data to only include relevant rows\n"
                           "- Remove unnecessary columns\n"
                           "- Split large files into smaller batches")
                    st.stop()
                else:
                    st.success(f"✅ Sheet '{selected_sheet_name}' is within limits and ready to process!")
            
            st.markdown("---")

        # ====================================================================
        # PROCESS BUTTON (only enabled if sheet selected and valid)
        # ====================================================================
        
        process_enabled = (
            st.session_state.selected_sheet is not None 
            and st.session_state.file_validated
        )
        
        if not process_enabled:
            st.warning("⚠️ Please select a valid sheet to continue")
        
        process_btn = render_process_button() if process_enabled else st.button("Process Offer", disabled=True, type="primary")

        if process_btn and process_enabled:
            with st.spinner("🔄 Processing your offer..."):
                success, output_path, df, error = process_uploaded_file(
                    uploaded_file=uploaded_file,
                    dept_type=dept_type,
                    double_stackable=st.session_state.double_stackable,
                    extract_price=st.session_state.extract_price,
                    product_images=None,
                    selected_sheet=st.session_state.selected_sheet,  # NEW: Pass selected sheet
                )

                if success:
                    df = _force_availability_ints(df)

                    st.session_state.processed = True
                    st.session_state.output_path = output_path
                    st.session_state.df = df
                    st.session_state.selected_df = None
                    st.session_state.product_images = {}
                    st.session_state.row_selected = None
                else:
                    st.error(f"❌ Error: {error}")

# ============================================================================
# RESULTS SECTION
# ============================================================================
if st.session_state.processed and st.session_state.df is not None:
    render_success_message()

    edited_df = render_selectable_table(st.session_state.df)

    if edited_df is not None and len(edited_df) > 0 and "Include" in edited_df.columns:
        selected_df = edited_df[edited_df["Include"] == True].copy()  # noqa: E712
        selected_df.drop(columns=["Include"], inplace=True, errors="ignore")
        selected_df.reset_index(drop=True, inplace=True)
    else:
        selected_df = edited_df.copy() if edited_df is not None else None
        if selected_df is not None:
            selected_df.reset_index(drop=True, inplace=True)

    if selected_df is not None:
        selected_df = _force_availability_ints(selected_df)

    st.session_state.selected_df = selected_df

    product_images = render_product_image_uploader(selected_df)
    st.session_state.product_images = product_images

    action, images_to_use = render_download_buttons(
        selected_df,
        product_images,
        dept_type=st.session_state.dept_type,
        base_filename=st.session_state.output_path.name,
    )

    # ------------------------------------------------------------------------
    # NO-IMAGES DOWNLOAD
    # ------------------------------------------------------------------------
    if action == "no_images" and selected_df is not None and len(selected_df) > 0:
        with st.spinner("📄 Generating Excel (no images)..."):
            from domain.schemas import FOOD_HEADERS, HPC_HEADERS
            from writers.excel_writer import write_rows_to_xlsx

            headers = FOOD_HEADERS if st.session_state.dept_type == "food" else HPC_HEADERS
            rows = selected_df.to_dict(orient="records")

            base = st.session_state.output_path.name if st.session_state.output_path else "offer.xlsx"
            output_no_images = Path(tempfile.gettempdir()) / f"no_images_{base}"

            write_rows_to_xlsx(
                output_path=output_no_images,
                sheet_name=st.session_state.dept_type.upper(),
                headers=headers,
                rows=rows,
                product_images=None,
            )

            st.success("✅ Excel (no images) generated!")

            with open(output_no_images, "rb") as f:
                st.download_button(
                    label="📥 Download Excel (No Images)",
                    data=f,
                    file_name=output_no_images.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="secondary",
                    width="stretch",
                    key="download_no_images",
                )

    # ------------------------------------------------------------------------
    # WITH-IMAGES DOWNLOAD
    # ------------------------------------------------------------------------
    if action == "with_images" and images_to_use and selected_df is not None and len(selected_df) > 0:
        with st.spinner("🎨 Generating Excel with images..."):
            from domain.schemas import FOOD_HEADERS, HPC_HEADERS
            from writers.excel_writer import write_rows_to_xlsx

            headers = FOOD_HEADERS if st.session_state.dept_type == "food" else HPC_HEADERS

            temp_dir = Path(tempfile.gettempdir()) / "offer_images"
            temp_dir.mkdir(exist_ok=True)

            image_paths: list[Path | None] = []
            for idx in range(len(selected_df)):
                if idx in images_to_use:
                    img_file = images_to_use[idx]
                    img_ext = Path(img_file.name).suffix
                    img_path = temp_dir / f"product_{idx}{img_ext}"

                    with open(img_path, "wb") as f:
                        f.write(img_file.getbuffer())

                    image_paths.append(img_path)
                else:
                    image_paths.append(None)

            rows = selected_df.to_dict(orient="records")

            base = st.session_state.output_path.name if st.session_state.output_path else "offer.xlsx"
            output_with_images = Path(tempfile.gettempdir()) / f"with_images_{base}"

            write_rows_to_xlsx(
                output_path=output_with_images,
                sheet_name=st.session_state.dept_type.upper(),
                headers=headers,
                rows=rows,
                product_images=image_paths,
            )

            st.success("✅ Excel with images generated!")

            with open(output_with_images, "rb") as f:
                st.download_button(
                    label="📥 Download Excel with Images",
                    data=f,
                    file_name=output_with_images.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    width="stretch",
                    key="final_download",
                )

    if render_reset_button():
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()
"""Read Excel files into raw dict format."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


def read_excel(xlsx_path: Path, sheet_name: str | None = None) -> List[Dict[str, Any]]:
    """Read Excel file where row 1 is headers, rows 2+ are data.
    
    Returns list of dicts with original supplier column names.
    No transformation, no interpretation - just raw data.
    
    Args:
        xlsx_path: Path to Excel file
        sheet_name: Optional sheet name (uses first sheet if None)
        
    Returns:
        List of row dicts with original headers as keys
    """
    xlsx_path = xlsx_path.expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    # Extract headers from row 1
    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        headers.append(str(h).strip() if h is not None else f"col_{c}")

    # Extract data rows
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row: Dict[str, Any] = {}
        is_empty = True
        
        for c, header in enumerate(headers, start=1):
            value = ws.cell(row=r, column=c).value
            if value not in (None, ""):
                is_empty = False
            row[header] = value
            
        if not is_empty:
            rows.append(row)

    wb.close()
    return rows
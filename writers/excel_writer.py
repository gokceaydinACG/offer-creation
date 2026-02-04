"""Excel writer with professional formatting - starts from B2 as required."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# Headers that must be forced into 2 lines (fixed regardless of zoom)
_FIXED_HEADER_LABELS = {
    "Piece per case": "Piece per\ncase",
    "Case per pallet": "Case per\npallet",
    "Pieces per pallet": "Pieces per\npallet",
    "Availability/Cartons": "Availability /\nCartons",
    "Availability/Pieces": "Availability /\nPieces",
    "Availability/Pallets": "Availability /\nPallets",
    "Price/Unit (Euro)": "Price / Unit\n(Euro)",
}

# Columns that should DISPLAY 1 decimal in Excel (value remains float)
_ONE_DECIMAL_COLUMNS = {
    "Availability/Cartons",
    "Availability/Pieces",
    "Availability/Pallets",
}


def _safe_float(v: Any) -> Optional[float]:
    """Best-effort convert to float; return None if not numeric."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip()
        if not s:
            return None
        s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None


def write_rows_to_xlsx(
    output_path: Path,
    sheet_name: str,
    headers: List[str],
    rows: List[Dict[str, Any]],
    product_images: Optional[List[Optional[Path]]] = None,
) -> None:
    """Write rows to Excel with professional formatting.

    Availability calculation (ONE-WAY):
    - Availability/Pieces: Editable by user (value)
    - Availability/Cartons: Formula (=Pieces ÷ Piece per case)
    - Availability/Pallets: Formula (=Pieces ÷ Pieces per pallet)
    
    When user edits Pieces → Cartons and Pallets auto-update via formulas.
    When user edits Cartons/Pallets → Formula is overwritten, no auto-update.
    """
    print(f"📝 Writing Excel (B2 start): {output_path.name}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Hide gridlines
    ws.sheet_view.showGridLines = False

    # Borders
    thin = Side(style="thin", color="000000")
    thick = Side(style="thick", color="000000")

    full_grid = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header border (NO vertical lines between columns)
    header_border_thin = Border(top=thin, bottom=thin)

    # Table starts at B2
    start_col = 2  # B
    start_row = 2  # row 2

    last_col = start_col + len(headers) - 1

    # Fixed header row height (tall enough for 2 lines)
    ws.row_dimensions[start_row].height = 32

    # Build header -> excel column index map
    header_to_excel_col: Dict[str, int] = {}
    for col_idx, header in enumerate(headers):
        header_to_excel_col[header] = start_col + col_idx

    # --- HEADERS ---
    for col_idx, header in enumerate(headers):
        excel_col = start_col + col_idx
        header_text = _FIXED_HEADER_LABELS.get(header, header)

        cell = ws.cell(row=start_row, column=excel_col, value=header_text)

        cell.font = Font(bold=True, color="000000", size=11, name="Calibri")
        cell.fill = PatternFill(start_color="FBF0D9", end_color="FBF0D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Base header border: only top+bottom thin (no vertical lines)
        cell.border = header_border_thin

    # Identify key columns for availability formulas (if present in this sheet)
    col_ppc = header_to_excel_col.get("Piece per case")
    col_ppp = header_to_excel_col.get("Pieces per pallet")
    col_pieces = header_to_excel_col.get("Availability/Pieces")
    col_cartons = header_to_excel_col.get("Availability/Cartons")
    col_pallets = header_to_excel_col.get("Availability/Pallets")

    # --- DATA ---
    for row_idx, row_data in enumerate(rows):
        excel_row = start_row + 1 + row_idx  # 3,4,5...

        # Write all cells as VALUES first
        for col_idx, header in enumerate(headers):
            excel_col = start_col + col_idx
            value = row_data.get(header)

            cell = ws.cell(row=excel_row, column=excel_col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = full_grid

            # Display 1 decimal for availability columns
            if header in _ONE_DECIMAL_COLUMNS and isinstance(value, (int, float)):
                cell.number_format = "0.0"
        
        # --- AUTOMATIC AVAILABILITY FORMULAS (ONE-WAY: Pieces → Cartons/Pallets) ---
        # Replace Cartons and Pallets values with formulas
        if col_pieces and col_cartons and col_pallets and col_ppc and col_ppp:
            pieces_ref = f"{get_column_letter(col_pieces)}{excel_row}"
            ppc_ref = f"{get_column_letter(col_ppc)}{excel_row}"
            ppp_ref = f"{get_column_letter(col_ppp)}{excel_row}"

            # Cartons = Pieces ÷ Piece per case
            cartons_cell = ws.cell(row=excel_row, column=col_cartons)
            cartons_cell.value = f'=IFERROR({pieces_ref}/{ppc_ref},"")'
            cartons_cell.number_format = "0.0"
            cartons_cell.alignment = Alignment(horizontal="center", vertical="center")
            cartons_cell.border = full_grid

            # Pallets = Pieces ÷ Pieces per pallet
            pallets_cell = ws.cell(row=excel_row, column=col_pallets)
            pallets_cell.value = f'=IFERROR({pieces_ref}/{ppp_ref},"")'
            pallets_cell.number_format = "0.0"
            pallets_cell.alignment = Alignment(horizontal="center", vertical="center")
            pallets_cell.border = full_grid

    # --- THICK BORDERS: ONLY TABLE TOP + HEADER ROW LEFT/RIGHT ---
    for col in range(start_col, last_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=thick,
            bottom=cell.border.bottom,
        )

    left_cell = ws.cell(row=start_row, column=start_col)
    left_cell.border = Border(
        left=thick,
        right=left_cell.border.right,
        top=left_cell.border.top,
        bottom=left_cell.border.bottom,
    )

    right_cell = ws.cell(row=start_row, column=last_col)
    right_cell.border = Border(
        left=right_cell.border.left,
        right=thick,
        top=right_cell.border.top,
        bottom=right_cell.border.bottom,
    )

    # --- TOTALS ROW FOR AVAILABILITY COLUMNS ---
    if rows and col_cartons and col_pieces and col_pallets:
        last_data_row = start_row + len(rows)
        totals_row = last_data_row + 1
        
        # Background color for totals row (availability columns only)
        totals_fill = PatternFill(start_color="F8F0D9", end_color="F8F0D9", fill_type="solid")
        
        # Border without top line (to avoid double border with data row above)
        totals_border = Border(left=thin, right=thin, top=None, bottom=thin)
        
        # Write totals only for the 3 availability columns
        for col_idx, header in enumerate(headers):
            excel_col = start_col + col_idx
            cell = ws.cell(row=totals_row, column=excel_col)
            
            if header == "Availability/Cartons":
                # SUM formula for cartons
                first_data_row = start_row + 1
                cell.value = f"=SUM({get_column_letter(excel_col)}{first_data_row}:{get_column_letter(excel_col)}{last_data_row})"
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = totals_fill  # Background color
                cell.border = totals_border  # No top border
                
            elif header == "Availability/Pieces":
                # SUM formula for pieces
                first_data_row = start_row + 1
                cell.value = f"=SUM({get_column_letter(excel_col)}{first_data_row}:{get_column_letter(excel_col)}{last_data_row})"
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = totals_fill  # Background color
                cell.border = totals_border  # No top border
                
            elif header == "Availability/Pallets":
                # SUM formula for pallets
                first_data_row = start_row + 1
                cell.value = f"=SUM({get_column_letter(excel_col)}{first_data_row}:{get_column_letter(excel_col)}{last_data_row})"
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = totals_fill  # Background color
                cell.border = totals_border  # No top border
            else:
                # Empty cell for other columns
                cell.value = None
                cell.border = totals_border  # No top border

    # --- TOTALS ROW FOR AVAILABILITY COLUMNS ---
    if col_cartons and col_pallets and len(rows) > 0:
        last_data_row = start_row + len(rows)
        totals_row = last_data_row + 1
        
        # Add empty cells with borders for all columns
        for col_idx, header in enumerate(headers):
            excel_col = start_col + col_idx
            cell = ws.cell(row=totals_row, column=excel_col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Determine borders (thick on left/right edges, top border thick)
            left_border = thick if excel_col == start_col else thin
            right_border = thick if excel_col == last_col else thin
            
            cell.border = Border(top=thick, bottom=thin, left=left_border, right=right_border)
        
        # Add totals for the 3 availability columns (will overwrite empty cells)
        first_data_row = start_row + 1
        
        # Availability/Cartons
        cartons_total_cell = ws.cell(row=totals_row, column=col_cartons)
        cartons_range = f"{get_column_letter(col_cartons)}{first_data_row}:{get_column_letter(col_cartons)}{last_data_row}"
        cartons_total_cell.value = f"=SUM({cartons_range})"
        cartons_total_cell.number_format = "0.0"
        cartons_total_cell.alignment = Alignment(horizontal="center", vertical="center")
        cartons_total_cell.font = Font(bold=True, size=10)
        # Keep existing border from loop above
        
        # Availability/Pieces
        pieces_total_cell = ws.cell(row=totals_row, column=col_pieces)
        pieces_range = f"{get_column_letter(col_pieces)}{first_data_row}:{get_column_letter(col_pieces)}{last_data_row}"
        pieces_total_cell.value = f"=SUM({pieces_range})"
        pieces_total_cell.number_format = "0.0"
        pieces_total_cell.alignment = Alignment(horizontal="center", vertical="center")
        pieces_total_cell.font = Font(bold=True, size=10)
        # Keep existing border from loop above
        
        # Availability/Pallets
        pallets_total_cell = ws.cell(row=totals_row, column=col_pallets)
        pallets_range = f"{get_column_letter(col_pallets)}{first_data_row}:{get_column_letter(col_pallets)}{last_data_row}"
        pallets_total_cell.value = f"=SUM({pallets_range})"
        pallets_total_cell.number_format = "0.0"
        pallets_total_cell.alignment = Alignment(horizontal="center", vertical="center")
        pallets_total_cell.font = Font(bold=True, size=10)
        # Keep existing border from loop above

    # --- OPTIMIZED COLUMN WIDTHS (NARROWER FOR BETTER PAGE FIT) ---
    # Define fixed widths for specific columns to fit page better
    COLUMN_WIDTHS = {
        "Article Number": 13,
        "EAN code unit": 15,
        "Product Description": 35,
        "Content": 10,
        "Languages": 25,
        "Piece per case": 11,
        "Case per pallet": 11,
        "Pieces per pallet": 12,
        "BBD": 12,
        "Availability/Cartons": 13,
        "Availability/Pieces": 13,
        "Availability/Pallets": 13,
        "Price/Unit (Euro)": 12,
    }
    
    for col_idx, header in enumerate(headers):
        excel_col = start_col + col_idx
        column_letter = get_column_letter(excel_col)
        
        # Use fixed width if defined, otherwise auto-calculate
        if header in COLUMN_WIDTHS:
            width = COLUMN_WIDTHS[header]
        else:
            # Fallback: auto-calculate for any undefined columns
            header_for_len = str(_FIXED_HEADER_LABELS.get(header, header)).replace("\n", " ")
            max_length = len(header_for_len)
            for row_data in rows:
                v = row_data.get(header)
                if v is not None:
                    max_length = max(max_length, len(str(v)))
            width = min(max_length + 3, 50)
        
        ws.column_dimensions[column_letter].width = width

    # No freeze panes
    ws.freeze_panes = None

    # --- IMAGES BELOW TABLE ---
    if product_images:
        valid_images = [p for p in product_images if p and Path(p).exists()]
        print(f"📸 Adding {len(valid_images)} product images...")

        # Calculate last row (data rows + totals row if exists)
        last_row = start_row + len(rows)
        if rows and col_cartons and col_pieces and col_pallets:
            last_row += 1  # Add 1 for totals row
        
        image_start_row = last_row + 5
        images_per_row = 4
        col_spacing = 3
        row_spacing = 9

        for img_idx, img_path in enumerate(product_images):
            if not img_path or not Path(img_path).exists():
                continue

            try:
                grid_row = img_idx // images_per_row
                grid_col = img_idx % images_per_row

                excel_row = image_start_row + (grid_row * row_spacing)
                excel_col = start_col + (grid_col * col_spacing)

                ws.row_dimensions[excel_row].height = 115

                img = XLImage(str(img_path))

                if img.width > img.height:
                    img.width = 150
                    img.height = int(150 * img.height / img.width)
                else:
                    img.height = 150
                    img.width = int(150 * img.width / img.height)

                cell_ref = f"{get_column_letter(excel_col)}{excel_row}"
                ws.add_image(img, cell_ref)

            except Exception as e:
                print(f"⚠️  Image error: {e}")

        print(f"✅ Added {len(valid_images)} images")

    wb.save(output_path)
    print(f"✅ Excel saved: {output_path.name}")
    
    # Log automatic formula info
    if col_pieces and col_cartons and col_pallets:
        print(f"ℹ️  Availability: Edit 'Pieces' column → Cartons & Pallets auto-update")